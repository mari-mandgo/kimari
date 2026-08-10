# デモ用の当初見積書（Excel）を作る。
#
# 実在の単価表は使わない。項目名は業界共通のもの、単価は架空。
# 「契約に何が含まれているか」を読み取る機能の実演に使うためのもので、
# 金額そのものはデモに関係しない。
#
#   python make-demo-estimate.py <出力パス>

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 工種ごとの項目。単価は架空の丸い数字にしてある
SECTIONS: list[tuple[str, list[tuple[str, str, str, float, int]]]] = [
    (
        "仮設工事",
        [
            ("現地調査費", "", "式", 1, 30000),
            ("工事申請代行・近隣調整費", "管理組合提出", "式", 1, 45000),
            ("共用部仮設養生費", "廊下・EV・階段", "式", 1, 60000),
            ("内部養生費", "", "㎡", 85, 500),
            ("発生廃材処分運搬費", "", "回", 4, 35000),
            ("清掃・片付費", "", "式", 1, 40000),
            ("資材運搬費", "荷揚げ共", "式", 1, 55000),
            ("美装工事", "内部クリーニング", "式", 1, 65000),
        ],
    ),
    (
        "解体工事",
        [
            ("内部解体工事", "スケルトン", "㎡", 85, 4500),
            ("解体廃材処分費", "", "㎡", 85, 2800),
            ("解体廃材運搬費", "", "㎡", 85, 900),
        ],
    ),
    (
        "木工事",
        [
            ("置床工事", "フリーフロア 際根太共", "㎡", 78, 6800),
            ("捨貼り合板", "t=12mm", "㎡", 78, 1800),
            ("無垢フローリング材", "ナラ オイルクリア", "㎡", 78, 9800),
            ("フローリング貼手間", "", "㎡", 78, 3200),
            ("間仕切壁下地組", "", "㎡", 46, 4200),
            ("間仕切壁PB貼", "t=12.5mm", "㎡", 46, 1600),
            ("壁 耐水PB貼", "洗面室", "㎡", 12, 2100),
            ("下地補強", "開口部共", "式", 1, 48000),
            ("天井下地組", "", "㎡", 78, 3800),
            ("天井PB貼", "t=9.5mm", "㎡", 78, 1500),
            ("玄関框・巾木取付", "ナラ材", "m", 32, 3400),
            ("金物費・雑材費", "", "㎡", 85, 800),
        ],
    ),
    (
        "建具工事",
        [
            ("室内ドア 開き戸", "既製品 枠共", "本", 4, 68000),
            ("上記吊込・調整費", "", "式", 1, 32000),
            ("造作建具枠材料費", "", "m", 24, 4200),
        ],
    ),
    (
        "内装工事",
        [
            ("壁・天井 クロス貼", "量産品", "㎡", 246, 1250),
            ("床 塩ビタイル貼", "洗面・トイレ", "式", 1, 68000),
            ("巾木・見切取付", "", "m", 62, 900),
        ],
    ),
    (
        "塗装工事",
        [
            ("壁 AEP塗装", "パテ処理共", "㎡", 48, 2200),
            ("出隅コーナー補強", "", "式", 1, 18000),
            ("玄関土間 塗装", "", "式", 1, 42000),
        ],
    ),
    (
        "住設工事",
        [
            ("システムキッチン本体", "I型 W2550", "セット", 1, 620000),
            ("キッチン 事前調査・組立費", "", "式", 1, 85000),
            ("キッチン 運搬・配送費", "", "式", 1, 38000),
            ("レンジフード", "", "台", 1, 78000),
            ("ユニットバス本体", "1620サイズ", "セット", 1, 780000),
            ("ユニットバス 組立費", "", "式", 1, 145000),
            ("洗面化粧台セット", "W750", "セット", 1, 168000),
            ("トイレ", "ウォシュレット一体型", "セット", 1, 195000),
            ("給湯器", "20号 追い炊き有", "台", 1, 235000),
        ],
    ),
    (
        "電気設備工事",
        [
            ("既設機器・配線撤去工事", "", "式", 1, 55000),
            ("配線工事", "材工", "式", 1, 285000),
            ("スイッチ", "材工", "組", 18, 4800),
            ("コンセント", "ダブル 材工", "組", 24, 5200),
            ("分電盤設置工事", "新規", "式", 1, 92000),
            ("照明器具取付工事", "", "式", 1, 68000),
            ("換気扇・ダクト工事", "材工", "式", 1, 125000),
            ("インターホン移設工事", "既存オートロック連動", "式", 1, 48000),
        ],
    ),
    (
        "給排水衛生設備工事",
        [
            ("止水工", "既設撤去・水抜き", "式", 1, 38000),
            ("給水給湯管", "架橋ポリエチレン管 13A", "m", 42, 3800),
            ("排水管", "塩ビ管 VP75A・VP50A", "m", 28, 4200),
            ("継手類・支持金物", "", "式", 1, 68000),
            ("配管保温", "", "m", 42, 1200),
            ("配管工事", "", "式", 1, 220000),
            ("水圧試験・通水試験", "", "式", 1, 32000),
            ("機器取付工事", "", "式", 1, 165000),
        ],
    ),
    (
        "ガス設備工事",
        [
            ("既設配管撤去", "", "式", 1, 28000),
            ("フレキ配管工事", "材工", "式", 1, 78000),
            ("給湯器取付", "", "式", 1, 52000),
            ("コンロ設置工事", "", "式", 1, 32000),
        ],
    ),
    (
        "諸経費",
        [
            ("設計料", "工事費の10%", "式", 1, 720000),
            ("現場管理費", "工事費の10%", "式", 1, 720000),
            ("諸経費", "一般管理費・会社経費", "式", 1, 360000),
        ],
    ),
]

NOTES = [
    "【備考】",
    "本見積書に含まれていない項目は別途工事といたします。",
    "共用部（サッシ・バルコニー・立管）の工事は含まれておりません。",
    "解体前に予測ができない事柄で追加・変更が発生した場合は別途となります。",
    "近隣、諸官庁、マンション管理組合等の指導による追加・変更は別途となります。",
    "タイル工事、造作浴室の防水工事は標準仕様には含まれません。",
    "仮設電力・仮設用水はお施主様のご支給をお願いいたします。",
]


def main() -> int:
    out = sys.argv[1] if len(sys.argv) > 1 else "当初見積書.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "御見積書"

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="E8E8E8")
    sec_fill = PatternFill("solid", fgColor="F5F5F5")

    ws["A1"] = "御 見 積 書"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A3"] = "工事件名"
    ws["B3"] = "田中様邸 マンションリノベーション工事"
    ws["A4"] = "工事場所"
    ws["B4"] = "東京都渋谷区◯◯町1-2-3"
    ws["A5"] = "工事概要"
    ws["B5"] = "専有部リノベーション工事（85.42㎡）"
    ws["A6"] = "工期"
    ws["B6"] = "協議の上決定"
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(bold=True)

    row = 8
    headers = ["工種", "名称", "規格・仕様", "単位", "数量", "単価", "金額", "備考"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    row += 1

    grand = 0
    for section, items in SECTIONS:
        start = row
        for i, (name, spec, unit, qty, price) in enumerate(items):
            amount = qty * price
            grand += amount
            values = [section if i == 0 else "", name, spec, unit, qty, price, amount, ""]
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = border
                if c in (5, 6, 7):
                    cell.number_format = "#,##0"
            row += 1
        for c in range(1, 9):
            ws.cell(row=start, column=c).fill = sec_fill

    ws.cell(row=row + 1, column=6, value="小計").font = Font(bold=True)
    ws.cell(row=row + 1, column=7, value=grand).number_format = "#,##0"
    ws.cell(row=row + 2, column=6, value="消費税(10%)")
    ws.cell(row=row + 2, column=7, value=int(grand * 0.1)).number_format = "#,##0"
    ws.cell(row=row + 3, column=6, value="合計").font = Font(bold=True, size=12)
    total = ws.cell(row=row + 3, column=7, value=grand + int(grand * 0.1))
    total.number_format = "#,##0"
    total.font = Font(bold=True, size=12)

    row += 5
    for line in NOTES:
        ws.cell(row=row, column=1, value=line)
        row += 1

    for col, width in zip("ABCDEFGH", [16, 34, 26, 8, 8, 12, 14, 20]):
        ws.column_dimensions[col].width = width

    wb.save(out)
    print(f"{out} を作成しました（税抜 {grand:,} 円 / 税込 {grand + int(grand * 0.1):,} 円）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
